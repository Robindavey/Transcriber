import openpyxl
import os

def xlsx_to_text(xlsx_path: str, txt_output_path: str = None) -> str:
    """
    Extracts text from an XLSX file.

    Args:
        xlsx_path: Path to the input XLSX
        txt_output_path: Optional path for output .txt file.
                         If None, will create next to XLSX.

    Returns:
        The extracted text as a string.
    """
    if txt_output_path is None:
        txt_output_path = os.path.splitext(xlsx_path)[0] + ".txt"

    wb = openpyxl.load_workbook(xlsx_path)
    full_text = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        full_text.append(f"Sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            full_text.append(row_text)
        full_text.append("")  # Blank line between sheets

    extracted_text = "\n".join(full_text)

    with open(txt_output_path, "w", encoding="utf-8") as f:
        f.write(extracted_text)

    return extracted_text